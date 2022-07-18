#! /usr/bin/env python3

from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment

from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from .models import Factura, Proveedor, ItemFactura, NotaDeCredito, NotaDeDebito, Pagos, InstrumentoPago
from mercaderias.models import CartaDePorte, CartaDePorteDescargada
excel_document = openpyxl.load_workbook('/home/hernan/Documentos/proyectos/adminweb/facturas.xlsx')
excel_prov = openpyxl.load_workbook('/home/hernan/Documentos/proyectos/adminweb/proveedores.xlsx')
excel_pagos = openpyxl.load_workbook('/home/hernan/Documentos/proyectos/adminweb/pagos.xlsx')
excel_parte = openpyxl.load_workbook('/media/hernan/MEMORIA USB/Aeq/Escritorio/21-22/PARTE COSECHA.xlsx')

def dataProv():
    print('-'*100)   
    provs = excel_prov.get_sheet_by_name('proveedores')
    
    for row in provs.iter_rows(min_row=2, max_col=10, max_row=405, values_only=True):
        if row[1]:
            p = Proveedor()
            p.nombre = row[0].upper()
            p.cuit = int(row[1])
            try:
                p.save()
            except:
                print(f'no se pudo guardar el proveedor {p}')

def dataExcel(): 
    print('-'*100)   
    cptes = excel_document.get_sheet_by_name('comprobantes')
    conta = 0
    for row in cptes.iter_rows(min_row=2, max_col=12, max_row=245, values_only=True):
        if row[0] == 'FAC':
            prov = Proveedor.objects.filter(nombre__icontains=row[4])
            
            fact_obj = Factura()
            fact_obj.fecha = row[5]
            fact_obj.fecha_vto = row[5]
            fact_obj.numero = f'{row[1]} - {row[2]} - {row[3]}'
            fact_obj.proveedor = prov[0]
            fact_obj.impuesto = row[8]
            fact_obj.otro_importe = row[9]
            fact_obj.rubro = row[10]
            fact_obj.cosecha = row[11]
            fact_obj.save()
            it = ItemFactura()
            it.cantidad = 1
            it.precio_unitario = row[6]
            it.factura = fact_obj
            it.save()
        
        if row[0] == 'NCC' or row[0] == 'NCR':
            prov = Proveedor.objects.filter(nombre__icontains=row[4])          
            fact_obj = NotaDeCredito()
            fact_obj.fecha = row[5]
            fact_obj.fecha_vto = row[5]
            fact_obj.numero = f'{row[1]} - {row[2]} - {row[3]}'
            fact_obj.proveedor = prov[0]
            fact_obj.impuesto = row[8]
            fact_obj.otro_importe = row[9]
            fact_obj.rubro = row[10]
            fact_obj.cosecha = row[11]
            fact_obj.save()
            it = ItemFactura()
            it.cantidad = 1
            it.precio_unitario = row[6]
            it.factura = fact_obj
            it.save()

        if row[0] == 'NDE':
            prov = Proveedor.objects.filter(nombre__icontains=row[4])          
            fact_obj = NotaDeDebito()
            fact_obj.fecha = row[5]
            fact_obj.fecha_vto = row[5]
            fact_obj.numero = f'{row[1]} - {row[2]} - {row[3]}'
            fact_obj.proveedor = prov[0]
            fact_obj.impuesto = row[8]
            fact_obj.otro_importe = row[9]
            fact_obj.rubro = row[10]
            fact_obj.cosecha = row[11]
            fact_obj.save()
            it = ItemFactura()
            it.cantidad = 1
            it.precio_unitario = row[6]
            it.factura = fact_obj
            it.save()

def deleteFacturas():
    fcs = Factura.objects.all()
    itms = ItemFactura.objects.all()
    
    for itm in itms:
        itm.delete()
    for fc in fcs:
        fc.delete()
            
def dataPagos():
    pagos_sheet = excel_pagos.get_sheet_by_name('pagos')    
    for row in pagos_sheet.iter_rows(min_row=2, max_col=10, max_row=217, values_only=True):    
        proveedor_hoja = row[2]       
        fecha_pago = datetime.strptime(row[3], '%d/%m/%Y %H:%M:%S')
        try:
            proveedor_obj = Proveedor.objects.get(nombre__endswith=proveedor_hoja)
            ipago_obj = InstrumentoPago()
            ipago_obj.fecha = fecha_pago.date()
            ipago_obj.instrumento = 'TRANSFERENCIA'
            ipago_obj.importe = row[4]
            ipago_obj.estado_uilizado = True
            ipago_obj.proveedor = proveedor_obj
            ipago_obj.save()
            pago_obj = Pagos()
            pago_obj.proveedor = proveedor_obj
            pago_obj.fecha = fecha_pago.date()
            pago_obj.numero = f'{row[0]}-{row[1]}'
            pago_obj.save()
            pago_obj.instrumentos.add(ipago_obj)
        except MultipleObjectsReturned:
            print(f'retorna mas de uno {proveedor_hoja}')
            
        except ObjectDoesNotExist:
            print(f'No existe el proveedor {proveedor_hoja}')
            
def delete_pagos():
    pagos = Pagos.objects.all() 
    for pago in pagos:
        pago.delete()
    ipagos = InstrumentoPago.objects.all()
    for ip in ipagos:
        ip.delete()       


def actualizarParte():
    print('actualizando parte cosecha')
    parte = excel_parte.get_sheet_by_name('PARTE')
    query= CartaDePorte.objects.filter(cosecha__iexact='2021/2022').filter(grano__iexact='MAIZ').order_by('numero')
    count=query.count()-1
    
    fuente = Font(name = 'DejaVu Sans', size = 14)
    centrado = Alignment(horizontal="center")
    
    '''for row in parte.iter_rows(min_row=2, max_col=20, values_only=False):
        for cell in row:
            cell.font = fuente
            cell.alignment = centrado

        if row[0].value == None:
            parte.cell(row = row[0].row, column = 1, value = query[count].fecha)
            parte.cell(row = row[0].row, column = 3, value = query[count].direccion_procedencia)
            parte.cell(row = row[0].row, column = 5, value = int(query[count].numero))
            parte.cell(row = row[0].row, column = 11, value = int(query[count].ctg))
            parte.cell(row = row[0].row, column = 12, value = query[count].contrato)
            parte.cell(row = row[0].row, column = 13, value = int(query[count].peso_neto))
            parte.cell(row = row[0].row, column = 21, value = float(query[count].tarifa))

            try:
                cp_desc = CartaDePorteDescargada.objects.get(cp__ctg__iexact=int(query[count].ctg))
                parte.cell(row = row[0].row, column = 14, value = int(cp_desc.kilos_descargados))
                parte.cell(row = row[0].row, column = 17, value = int(cp_desc.observaciones_descarga))
            except:
                pass
           
            count-=1
            if count<0:
                excel_parte.save('/media/hernan/MEMORIA USB/Aeq/Escritorio/21-22/PARTE COSECHA.xlsx')
                break
        else:
            try:
                cp_desc = CartaDePorteDescargada.objects.get(cp__ctg__iexact=str(row[10].value))
                parte.cell(row = row[0].row, column = 14, value = int(cp_desc.kilos_descargados))
                parte.cell(row = row[0].row, column = 17, value = int(cp_desc.observaciones_descarga))
                excel_parte.save('/media/hernan/MEMORIA USB/Aeq/Escritorio/21-22/PARTE COSECHA.xlsx')
            except:
                pass
    '''
    

        

        
        


        

 




        


