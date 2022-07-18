#! /usr/bin/env python3

import openpyxl
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from .models import CartaDePorte, CartaDePorteDescargada, Destinatario,\
 Destino, RemitenteComercial,ContratoMercaderia

excel_document = openpyxl.load_workbook('/home/hernan/Documentos/proyectos/adminweb/parte2021.xlsx')

def deleteDescargadas():
    cpd = CartaDePorteDescargada.objects.filter(cp__cosecha__icontains='2021/2022')
    for e in cpd:
        e.delete()

def actualizar():
    cps = CartaDePorte.objects.all()
    for cp in cps:
        if cp.direccion_procedencia == 'Del sol':
            cp.direccion_procedencia = 'CORRAL DE BUSTOS'

        if cp.direccion_procedencia == 'Vidaurre':
            cp.direccion_procedencia = 'VIDAURRON'

        cp.establecimiento = cp.direccion_procedencia
        
        print(cp.direccion_procedencia)
        #cp.save()

def parseCp():
    cps = CartaDePorte.objects.all()
    for cp in cps:
        cp.direccion_procedencia = 'ZONA RURAL S/N'
        cp.save()


def establecim():
    parte = excel_document.get_sheet_by_name('PARTE')
    count=0
    for row in parte.iter_rows(min_row=2, max_col=21, max_row=500, values_only=True):
        if row[4]:
            try:
                cp_obj = CartaDePorte.objects.get(numero__iexact=int(row[4]))
                cp_obj.establecimiento = row[2]
                #print(cp_obj.establecimiento)
                cp_obj.save()
            except:
                count+=1
                pass
    print(f'cant filas no agregadas {count}')

def dataExcel():    
    print('-'*50)    
    parte = excel_document.get_sheet_by_name('pte')
    listancp = []
    noenlista = []
    total_cp = 0
    for row in parte.iter_rows(min_row=2, max_col=21, max_row=200, values_only=True):   
        if row[4] != None and int(row[4])>103:
            n_cp = CartaDePorte()
            n_cp.fecha = row[0]
            n_cp.localidad_procedencia = row[1]
            n_cp.direccion_procedencia = row[2]
            n_cp.numero = int(row[4])
            print(row[7])
            dest = Destinatario.objects.get(nombre__startswith=row[7])
            n_cp.destinatario = dest            
            print(row[9])
            destno = Destino.objects.get(nombre__endswith=row[9])            
            n_cp.destino = destno

            try:
                rem_c = RemitenteComercial.objects.get(nombre__endswith=row[8])
                n_cp.remitente_comercial = rem_c
            except:
                pass
            n_cp.ctg = row[10]          
            n_cp.peso_neto = row[12]
            n_cp.grano = row[18]
            if row[13]>0:
                n_cpd = CartaDePorteDescargada()
                n_cpd.cp = n_cp                
                n_cpd.kilos_descargados = int(row[13])
                try:
                    ctto = ContratoMercaderia.objects.get(nro_vendedor__iexact=row[11])
                    n_cpd.contrato = ctto
                except:
                    pass
                if row[15]:                    
                    n_cpd.kilos_merma = row[15]
                if row[16]:
                    n_cpd.observaciones_descarga = row[16]
                n_cp.save()
                n_cpd.save()
                total_cp+=1               
            else:
                n_cp.save()            
                total_cp+=1
    print('Total cps: ',total_cp)

    '''

        n_cp = CartaDePorteDescargada()
        n_cp.kilos_descargados = row[13]   
        if row[15]:
            n_cp.kilos_merma = row[15]
        else:
            n_cp.kilos_merma = 0
        n_cp.observaciones_descarga = ''    
        
        try:
            cp = CartaDePorte.objects.get(numero=row[4])    
            cp.grano = row[18].upper()
            cp.direccion_procedencia = row[2].upper()  
            cp.save()
            n_cp.cp = cp
            n_cp.save()

        except ObjectDoesNotExist:
            carta = CartaDePorte()
            carta.fecha = row[0]
            carta.numero = row[4]
            carta.ctg = row[10]
            carta.grano = row[18].upper()
            carta.cosecha = '2021/2022'
            carta.peso_neto = row[12]
            carta.direccion_procedencia = row[2].upper()
            carta.localidad_procedencia = row[1].upper()    
            carta.save()
            n_cp.cp = carta
            n_cp.save()

        except MultipleObjectsReturned:            
            cps = CartaDePorte.objects.filter(numero=row[4])
            for cp in cps:
                if int(cp.peso_neto) == row[12] or int(cp.peso_neto) == row[13]:
                    cp.direccion_procedencia = row[2].upper()
                    cp.save()
                    n_cp.cp = cp
                    n_cp.save()

        except Exception as e:
            noenlista.append(row[4])
            total_cp+=1
    print(f'total cp: { total_cp }')        
    print('total cp no parseadas: ', noenlista)

'''